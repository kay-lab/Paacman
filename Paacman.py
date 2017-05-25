#! /usr/bin/env python

#Protein amino acid composition analysis = Paacman
#Version 1.0 (released 5/22/2017 by Patrick Erickson)

#Please visit the Paacman Github repository to view the legal license before using
#Paacman: https://github.com/kay-lab/Paacman

#This python script will analyze the .txt files containing desired proteins'
#FASTA info in order to analyze each protein's amino acid composition. This will
#analyze the amino acid composition of an entire group of proteins that are
#within the same folder. The script also analyzes di-AA sequences for all proteins.

#Import important modules.
import re
#This allows FASTA .txt files to be read in numerical and alphabetical order by file name
numbers = re.compile(r'(\d+)')
import openpyxl #Needs to be installed by the user!
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from string import ascii_uppercase
import glob
from collections import Counter
import os
import sys

#This allows for the proteins to be labeled in numerical order.
def numericalSort(value):
	"""Function for sorting files in numerical and alphabetical ascending order"""
	parts = numbers.split(value)
	parts[1::2] = map(int, parts[1::2])
	return parts
    
#This allows overlapping di-AA sequences (e.g., EEE = 2 EE's) to be counted correctly.
def occurrences(string, sub):
	"""Function for counting the number of di-AA sequences that overlap (e.g., AAA)"""
	count = start = 0
	while True:
		start = string.find(sub, start) + 1
		if start > 0:
			count+=1
		else:
			return count

#Creates colors to fill in Excel cells (openpyxl).
blackFill = PatternFill(start_color='FF000000',
                        end_color='FF000000',
                        fill_type='solid')

redFill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')

aquaFill = PatternFill(start_color='007FFFD4',
                       end_color='007FFFD4',
                       fill_type='solid')

greenFill = PatternFill(start_color='FF00FF00',
                       end_color='FF00FF00',
                       fill_type='solid')

heatBlueFill = PatternFill(start_color='FF0000AA',
                       end_color='FF0000AA',
                       fill_type='solid')

heatWhiteFill = PatternFill(start_color='FFFFFFFF',
                       end_color='FFFFFFFF',
                       fill_type='solid')

heatRedFill = PatternFill(start_color='FFAA0000',
                       end_color='FFAA0000',
                       fill_type='solid')

#Setting for a thick border line (openpyxl).
thick = Side(style='thick')

#Setting for centering a cell (openpyxl).
center = Alignment(horizontal="center", vertical="center")


#List of canonical amino acids.
AAList = ["A","C","D","E","F","G","H","I","K","L",
          "M","N","P","Q","R","S","T","V","W","Y"]
          
#List of Cys ligation sites.
CysLigList = ["AC","CC","DC","EC","FC","GC","HC","IC","KC","LC",
          	  "MC","NC","PC","QC","RC","SC","TC","VC","WC","YC"]
          	  
#List of Ala ligation sites.
AlaLigList = ["AA","CA","DA","EA","FA","GA","HA","IA","KA","LA",
              "MA","NA","PA","QA","RA","SA","TA","VA","WA","YA"]

#List of Potential Aspartimides.
AspartimideList = ["DA","DC","DD","DE","DF","DG","DH","DI","DK","DL",
          		   "DM","DN","DP","DQ","DR","DS","DT","DV","DW","DY"]

#List of Potential Pseudoprolines
PSList = ["AS", "AT", "DS", "DT", "ES", "ET", "FS", "FT", "GS", "GT", "HS", "HT",
		  "IS", "IT", "KS", "KT", "LS", "LT", "MS", "MT", "NS", "NT", "QS", "QT",
		  "RS", "RT", "VS", "VT", "WS", "WT", "YS", "YT"]

#Creates List of all Possible Di-peptides.
AllDiPeptideList = []
for i in AAList:
	rowList = []
	for j in AAList:
		rowList.append(i + j)
	AllDiPeptideList.append(rowList)

#Intro to the user.
print "Welcome to Paacman! Starting amino acid composition analysis..."
print ""

#Creates the output Excel file with the desired 3 sheets.
outFile = openpyxl.Workbook()
outFile.create_sheet(index=0, title="AA Composition")
outFile.create_sheet(index=1, title="CPS Di-AA Composition")
outFile.create_sheet(index=2, title="Total Di-AA Composition")
outFile.remove_sheet(outFile.get_sheet_by_name("Sheet")) #Removes empty sheet

#The following is for the single AA Composition portion of Paacman.
#Writes initial information into the AA Composition sheet.
sheet = outFile.get_sheet_by_name("AA Composition")
sheet.merge_cells("B1:U1")
sheet['B1'] = "Amino Acids"
sheet['B1'].alignment = center
sheet['B1'].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet['B1'].fill = blackFill

sheet['A2'] = "Protein Name"
sheet['A2'].alignment = center
sheet['A2'].font = Font(size=12, bold=True)
sheet['A2'].border = Border(right=thick, bottom=thick, top=thick)

#This writes the amino acid single letter codes at the top of each column.
letterList = list(ascii_uppercase)
letterList.pop(0) #Skips the first cell since Protein Name is written there
AAEntryCount = 0
for i in AAList:
    cellNum = letterList[AAEntryCount] + "2"
    sheet[cellNum] = i
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].border = Border(bottom=thick)
    AAEntryCount += 1

#Types "Total" heading for last column.
sheet['V2'] = "Total"
sheet['V2'].alignment = center
sheet['V2'].font = Font(size=12, bold=True)
sheet['V2'].border = Border(right=thick, bottom=thick, top=thick, left=thick)

#Types AA counts for each protein into approp. cells in spreadsheet.
rowTracker = 3
firstDataRow = 3 #Important for AA percentage table and heat map
for File in sorted(glob.iglob("*.txt"), key=numericalSort):
    with open(File, 'U') as inFile: #U allows any .txt format to be accepted.
        ProteinName = File.rstrip(".txt")
        if ProteinName.endswith("fasta"): #Removes "fasta" from protein name, if there.
            ProteinName = ProteinName.rstrip(".fasta")
        sheet["A" + str(rowTracker)] = ProteinName
        sheet["A" + str(rowTracker)].alignment = center
        sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
        sheet["A" + str(rowTracker)].border = Border(right=thick)
        next(inFile) #Skips the info line in FASTA files
        
        #This stores the number of each amino acid in the protein.
        AACompDict = Counter(letter for line in inFile
                             for letter in line.upper()
                             if letter in ascii_uppercase or letter == ">")
    	
    	#This makes sure the user doesn't have multiple FASTA's in a single .txt file.
    	if ">" in AACompDict:
    		print "Your " + ProteinName + " file contains more than 1 FASTA file."
    		print "Paacman does not currently read .txt files with more than 1 FASTA."
    		print "Please reformat this file and try again."
    		print "Paacman terminated."
    		sys.exit()
        
        #This writes each amino acid count to the output Excel file.
        AAEntryCount = 0
        for i in AAList:
            cellNum = letterList[AAEntryCount] + str(rowTracker)
            sheet[cellNum] = AACompDict[i]
            sheet[cellNum].alignment = Alignment(horizontal="center")
            AAEntryCount +=1
        
        #This writes the total number of amino acids in the protein into the Excel sheet.
        rowTotal = ("=SUM(" + letterList[0] + str(rowTracker) + ":" +
                    letterList[AAEntryCount - 1] + str(rowTracker) + ")")
        sheet[letterList[AAEntryCount] + str(rowTracker)] = rowTotal
        sheet[letterList[AAEntryCount] + str(rowTracker)].alignment = center
        sheet[letterList[AAEntryCount] + str(rowTracker)].font = (
            Font(size=12, bold=True))
        sheet[letterList[AAEntryCount] + str(rowTracker)].fill = aquaFill
        sheet[letterList[AAEntryCount] + str(rowTracker)].border = (
            Border(left=thick, right=thick))
    rowTracker += 1
    lastDataRow = rowTracker #Important for AA percentage table and heat map
    inFile.close()

#Types "Total" heading into the row after all of the protein names.    
sheet["A" + str(rowTracker)] = "Total"
sheet["A" + str(rowTracker)].alignment = center
sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
sheet["A" + str(rowTracker)].border = Border(top=thick, right=thick, bottom=thick)

#Puts total AA count for each residue into the Excel sheet.
AAEntryCount = 0
for i in range(0,21):
    cellNum = letterList[AAEntryCount] + str(rowTracker)
    sheet[cellNum] = ("=SUM(" + letterList[AAEntryCount] + "3:" +
                      letterList[AAEntryCount] + str(rowTracker - 1) + ")")
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].fill = aquaFill
    sheet[cellNum].border = Border(top=thick, bottom=thick, right=thick)
    AAEntryCount += 1
    finalTotalCell = cellNum
rowTracker += 1

#Makes special formatting for the total AA count for the entire list of proteins.
sheet[finalTotalCell].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet[finalTotalCell].fill = redFill
sheet[finalTotalCell].border = Border(right=thick, left=thick, top=thick, bottom=thick)

#Puts 'Percentage' title for the row following the total number of amino acids row.
sheet["A" + str(rowTracker)] = "Percentage"
sheet["A" + str(rowTracker)].alignment = center
sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
sheet["A" + str(rowTracker)].border = Border(right=thick, bottom=thick)

#Enters percentage formulas into each cell in the 'Percentage' row.
AAEntryCount = 0
for i in range(0,20):
    cellNum = letterList[AAEntryCount] + str(rowTracker)
    sheet[cellNum] = ("=" + letterList[AAEntryCount] + str(rowTracker - 1) +
                      "/" + finalTotalCell)
    sheet[cellNum].number_format = '0.00%'
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].fill = greenFill
    sheet[cellNum].border = Border(bottom=thick, right=thick)
    AAEntryCount +=1

#The following codes for the AA Composition heat map.
#Writes the initial heat map information into the sheet.
rowTracker +=2 #Puts a row space between the heat map and total amino acid info.
sheet.merge_cells('B'+str(rowTracker)+":U"+str(rowTracker))
sheet['B'+str(rowTracker)] = "Amino Acid Percentage Heatmap"
sheet["B"+str(rowTracker)].alignment = center
sheet["B"+str(rowTracker)].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet["B"+str(rowTracker)].fill = blackFill

rowTracker +=1
sheet['A'+str(rowTracker)] = "Protein Name"
sheet['A'+str(rowTracker)].alignment = center
sheet['A'+str(rowTracker)].font = Font(size=12, bold=True)
sheet['A'+str(rowTracker)].border = Border(right=thick, bottom=thick, top=thick)

#Writes single letter amino acid codes into the top of the heat map.
AAEntryCount = 0
for i in AAList:
    cellNum = letterList[AAEntryCount] + str(rowTracker)
    sheet[cellNum] = i
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    if i == "Y": #For proper formatting of border in the heat map.
        sheet[cellNum].border = Border(bottom=thick, right=thick)
    else:
        sheet[cellNum].border = Border(bottom=thick)
    AAEntryCount += 1

#This saves the first row of the actual heat map, which helps write the heat map legend
#later on in the script.
rowTracker +=1
firstPercentageCell = 'B' + str(rowTracker)
firstPercentageRow = rowTracker

#This makes sure that the user has FASTA .txt files within their folder, as this is where
#the first error occurs if there are no files found in the folder.
try:
	lastDataRow
except NameError:
	print "There appears to be no FASTA .txt files in your folder!"
	print "Please make sure that your FASTA files are saved as .txt files."
	print "Paacman terminated"
	sys.exit()

#This loop writes the amino acid percentages into the heat map.
for i in range(firstDataRow,lastDataRow):
    #Writes the protein name into the heat map
    sheet['A'+str(rowTracker)] = '=A'+str(i)
    sheet['A'+str(rowTracker)].alignment = center
    sheet['A'+str(rowTracker)].font = Font(size=12, bold=True)
    sheet['A'+str(rowTracker)].border = Border(right=thick)
    
    #This writes the amino acid percentage (# of AA divided by total # of residues in a
    #protein) into the heat map.
    AAEntryCount = 0
    for letter in AAList:
        cellNum = letterList[AAEntryCount] + str(rowTracker)
        sheet[cellNum] = "=" + letterList[AAEntryCount] + str(i) + "/V" + str(i)
        sheet[cellNum].number_format = '0.00%' #Writes % and limits to 2 decimals
        sheet[cellNum].alignment = Alignment(horizontal="center")
        sheet[cellNum].font = Font(bold=True)
        if letter == "Y": #Important for proper heat map formatting
            sheet[cellNum].border = Border(right=thick)
        if i == lastDataRow-1:
            sheet[cellNum].border = Border(bottom=thick)
        AAEntryCount +=1
    rowTracker +=1
    
#This is for proper formatting of the heat map in the Excel sheet.
sheet['A'+str(rowTracker-1)].border = Border(right=thick, bottom=thick)
sheet['U'+str(rowTracker-1)].border = Border(right=thick, bottom=thick)

#This saves the last cell of the heat map (important for conditional formatting).
lastPercentageCell = 'U' + str(rowTracker-1)

#Setting for Conditional Formatting used in the AA heat map.
sheet.conditional_formatting.add(firstPercentageCell + ":" + lastPercentageCell,
                                 ColorScaleRule(start_type='num', start_value=0, start_color='FF0000AA',
                                 mid_type='num', mid_value=0.05, mid_color='FFFFFFFF',
                                 end_type='num', end_value=0.10, end_color='FFAA0000')
                                 )

#The following writes the heat map key into the Excel sheet.
sheet.merge_cells("W" + str(firstPercentageRow) + ":X" + str(firstPercentageRow))
sheet["W" + str(firstPercentageRow)] = "Heatmap Legend"
sheet["W" + str(firstPercentageRow)].alignment = Alignment(horizontal="center")
sheet["W" + str(firstPercentageRow)].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet["W" + str(firstPercentageRow)].fill = blackFill

sheet["W" + str(firstPercentageRow+1)] = "Color"
sheet["W" + str(firstPercentageRow+1)].alignment = Alignment(horizontal="center")
sheet["W" + str(firstPercentageRow+1)].font = Font(size=12, bold=True)
sheet["W" + str(firstPercentageRow+1)].border = Border(left=thick, right=thick, bottom=thick)

sheet["X" + str(firstPercentageRow+1)] = "%"
sheet["X" + str(firstPercentageRow+1)].alignment = Alignment(horizontal="center")
sheet["X" + str(firstPercentageRow+1)].font = Font(size=12, bold=True)
sheet["X" + str(firstPercentageRow+1)].border = Border(right=thick, bottom=thick)

sheet["W" + str(firstPercentageRow+2)].fill = heatBlueFill
sheet["W" + str(firstPercentageRow+2)].border = Border(right=thick, left=thick, bottom=thick)
sheet["W" + str(firstPercentageRow+3)].fill = heatWhiteFill
sheet["W" + str(firstPercentageRow+3)].border = Border(right=thick, left=thick, bottom=thick)
sheet["W" + str(firstPercentageRow+4)].fill = heatRedFill
sheet["W" + str(firstPercentageRow+4)].border = Border(right=thick, left=thick, bottom=thick)

sheet["X" + str(firstPercentageRow+2)] = "0.00%"
sheet["X" + str(firstPercentageRow+2)].alignment = Alignment(horizontal="center")
sheet["X" + str(firstPercentageRow+2)].font = Font(size=12, bold=True)
sheet["X" + str(firstPercentageRow+2)].border = Border(right=thick, bottom=thick)

sheet["X" + str(firstPercentageRow+3)] = "5.00%"
sheet["X" + str(firstPercentageRow+3)].alignment = Alignment(horizontal="center")
sheet["X" + str(firstPercentageRow+3)].font = Font(size=12, bold=True)
sheet["X" + str(firstPercentageRow+3)].border = Border(right=thick, bottom=thick)

sheet["X" + str(firstPercentageRow+4)] = "10.00%"
sheet["X" + str(firstPercentageRow+4)].alignment = Alignment(horizontal="center")
sheet["X" + str(firstPercentageRow+4)].font = Font(size=12, bold=True)
sheet["X" + str(firstPercentageRow+4)].border = Border(right=thick, bottom=thick)

#The following is for the complete protein synthesis (CPS) di-peptides in Paacman.
#Writes initial Cys ligation site information into the CPS Di-AA sheet.
sheet = outFile.get_sheet_by_name("CPS Di-AA Composition")
sheet.merge_cells("B1:U1")
sheet['B1'] = "Cysteine Ligation Sites"
sheet['B1'].alignment = center
sheet['B1'].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet['B1'].fill = blackFill

sheet['A2'] = "Protein Name"
sheet['A2'].alignment = center
sheet['A2'].font = Font(size=12, bold=True)
sheet['A2'].border = Border(right=thick, bottom=thick, top=thick)

#Writes the possible Cys ligation sites into the sheet.
letterList = list(ascii_uppercase)
letterList.pop(0) #skips the first cell since Protein Name is written there
CysLigEntryCount = 0
for i in CysLigList:
    cellNum = letterList[CysLigEntryCount] + "2"
    sheet[cellNum] = i
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].border = Border(bottom=thick)
    CysLigEntryCount += 1
    
#Types "Total" heading for last column.
sheet['V2'] = "Total"
sheet['V2'].alignment = center
sheet['V2'].font = Font(size=12, bold=True)
sheet['V2'].border = Border(right=thick, bottom=thick, top=thick, left=thick)

#Types Cys ligation site counts for each protein into approp. cells in sheet.
rowTracker = 3
for File in sorted(glob.iglob("*.txt"), key=numericalSort):
    with open(File, "U") as inFile: #U allows any format for .txt files
        #Writes protein name into sheet
        ProteinName = File.rstrip(".txt")
        if ProteinName.endswith("fasta"): #Removes "fasta" from protein name, if there.
            ProteinName = ProteinName.rstrip(".fasta")
        sheet["A" + str(rowTracker)] = ProteinName
        sheet["A" + str(rowTracker)].alignment = center
        sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
        sheet["A" + str(rowTracker)].border = Border(right=thick)
        inFile.readline() #Skips the first lines in FASTA file
        
        #This saves the entire protein sequence within a single variable,
        #allowing di-peptide searching.
        ProteinRead = inFile.read()
        ProteinRead = ProteinRead.replace("\n","").replace("\r","").replace(" ","").replace("\t","")
        ProteinRead = ProteinRead.upper() #Allows case-insensitivity
        
        #Writes the number of each Cys ligation site found within the sheet.
        CysLigEntryCount = 0
        for i in CysLigList:
            cellNum = letterList[CysLigEntryCount] + str(rowTracker)
            sheet[cellNum] = occurrences(ProteinRead, i)
            sheet[cellNum].alignment = Alignment(horizontal="center")
            CysLigEntryCount +=1
         
        #Writes total # of Cys ligation sites found within a protein into the sheet.   
        rowTotal = ("=SUM(" + letterList[0] + str(rowTracker) + ":" +
                    letterList[CysLigEntryCount - 1] + str(rowTracker) + ")")
        sheet[letterList[CysLigEntryCount] + str(rowTracker)] = rowTotal
        sheet[letterList[CysLigEntryCount] + str(rowTracker)].alignment = center
        sheet[letterList[CysLigEntryCount] + str(rowTracker)].font = (
            Font(size=12, bold=True))
        sheet[letterList[CysLigEntryCount] + str(rowTracker)].fill = aquaFill
        sheet[letterList[CysLigEntryCount] + str(rowTracker)].border = (
            Border(left=thick, right=thick))
    rowTracker += 1
    inFile.close()
    
#Types "Total" heading into the row after the final protein row. 
sheet["A" + str(rowTracker)] = "Total"
sheet["A" + str(rowTracker)].alignment = center
sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
sheet["A" + str(rowTracker)].border = Border(top=thick, right=thick, bottom=thick)

#Puts total Cys ligation count for each type of Cys ligation into the sheet.
CysLigEntryCount = 0
for i in range(0,21):
    cellNum = letterList[CysLigEntryCount] + str(rowTracker)
    sheet[cellNum] = ("=SUM(" + letterList[CysLigEntryCount] + "3:" +
                      letterList[CysLigEntryCount] + str(rowTracker - 1) + ")")
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].fill = aquaFill
    sheet[cellNum].border = Border(top=thick, bottom=thick, right=thick)
    CysLigEntryCount += 1
rowTracker += 2 #Skips space to enter the Ala ligation site information

#Writes initial Ala ligation site information into the CPS Di-AA sheet.
sheet.merge_cells("B" + str(rowTracker) + ":U" + str(rowTracker))
sheet['B' + str(rowTracker)] = "Alanine Ligation Sites"
sheet['B' + str(rowTracker)].alignment = center
sheet['B' + str(rowTracker)].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet['B' + str(rowTracker)].fill = blackFill
rowTracker += 1

sheet['A' + str(rowTracker)] = "Protein Name"
sheet['A' + str(rowTracker)].alignment = center
sheet['A' + str(rowTracker)].font = Font(size=12, bold=True)
sheet['A' + str(rowTracker)].border = Border(right=thick, bottom=thick, top=thick)

#This writes the possible Ala ligation sites into the top of each column.
letterList = list(ascii_uppercase)
letterList.pop(0) #Skips the first cell since Protein Name is written there
AlaLigEntryCount = 0
for i in AlaLigList:
    cellNum = letterList[AlaLigEntryCount] + str(rowTracker)
    sheet[cellNum] = i
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].border = Border(bottom=thick)
    AlaLigEntryCount += 1
    
#Types "Total" heading for the last column.
sheet['V' + str(rowTracker)] = "Total"
sheet['V' + str(rowTracker)].alignment = center
sheet['V' + str(rowTracker)].font = Font(size=12, bold=True)
sheet['V' + str(rowTracker)].border = Border(right=thick, bottom=thick, top=thick, left=thick)

#Types Ala ligation site counts for each protein into approp. cells in sheet.
rowTracker +=1
firstAlaRow = rowTracker #Important for getting the sum of each Ala junction later in script.
for File in sorted(glob.iglob("*.txt"), key=numericalSort):
    with open(File, "U") as inFile: #U allows any format for .txt files
        #Writes protein name into sheet.
        ProteinName = File.rstrip(".txt")
        if ProteinName.endswith("fasta"): #Removes fasta from protein name, if there
            ProteinName = ProteinName.rstrip(".fasta")
        sheet["A" + str(rowTracker)] = ProteinName
        sheet["A" + str(rowTracker)].alignment = center
        sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
        sheet["A" + str(rowTracker)].border = Border(right=thick)
        inFile.readline() #Skips the first line of a FASTA file
        
        #This saves the entire protein sequence within a single variable,
        #allowing di-peptide searching.
        ProteinRead = inFile.read()
        ProteinRead = ProteinRead.replace("\n","").replace("\r","").replace(" ","").replace("\t","")
        ProteinRead = ProteinRead.upper()
        
        #This writes the number of each type of Ala ligation site into the sheet.
        AlaLigEntryCount = 0
        for i in AlaLigList:
            cellNum = letterList[AlaLigEntryCount] + str(rowTracker)
            sheet[cellNum] = occurrences(ProteinRead, i)
            sheet[cellNum].alignment = Alignment(horizontal="center")
            AlaLigEntryCount +=1
        
        #This writes the total number of Ala ligation sites for a protein into the sheet.    
        rowTotal = ("=SUM(" + letterList[0] + str(rowTracker) + ":" +
                    letterList[AlaLigEntryCount - 1] + str(rowTracker) + ")")
        sheet[letterList[AlaLigEntryCount] + str(rowTracker)] = rowTotal
        sheet[letterList[AlaLigEntryCount] + str(rowTracker)].alignment = center
        sheet[letterList[AlaLigEntryCount] + str(rowTracker)].font = (
            Font(size=12, bold=True))
        sheet[letterList[AlaLigEntryCount] + str(rowTracker)].fill = aquaFill
        sheet[letterList[AlaLigEntryCount] + str(rowTracker)].border = (
            Border(left=thick, right=thick))
    rowTracker += 1
    inFile.close()
    
#Types "Total" heading into the row after the last protein row.
sheet["A" + str(rowTracker)] = "Total"
sheet["A" + str(rowTracker)].alignment = center
sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
sheet["A" + str(rowTracker)].border = Border(top=thick, right=thick, bottom=thick)

#Puts total Ala ligation count for each type of Ala junction into the sheet.
AlaLigEntryCount = 0
for i in range(0,21):
    cellNum = letterList[AlaLigEntryCount] + str(rowTracker)
    sheet[cellNum] = ("=SUM(" + letterList[AlaLigEntryCount] + str(firstAlaRow) + ":" +
                      letterList[AlaLigEntryCount] + str(rowTracker - 1) + ")")
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].fill = aquaFill
    sheet[cellNum].border = Border(top=thick, bottom=thick, right=thick)
    AlaLigEntryCount += 1
rowTracker += 2 #Skips a space to write in the possible aspartimides into the sheet

#Writes initial aspartimide information into the CPS Di-AA sheet.
sheet.merge_cells("B" + str(rowTracker) + ":U" + str(rowTracker))
sheet['B' + str(rowTracker)] = "Possible Aspartimides"
sheet['B' + str(rowTracker)].alignment = center
sheet['B' + str(rowTracker)].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet['B' + str(rowTracker)].fill = blackFill
rowTracker += 1

sheet['A' + str(rowTracker)] = "Protein Name"
sheet['A' + str(rowTracker)].alignment = center
sheet['A' + str(rowTracker)].font = Font(size=12, bold=True)
sheet['A' + str(rowTracker)].border = Border(right=thick, bottom=thick, top=thick)

#Writes the possible aspartimides into the top of each column.
letterList = list(ascii_uppercase)
letterList.pop(0) #skips the first cell since Protein Name is written there
AspartimideEntryCount = 0
for i in AspartimideList:
    cellNum = letterList[AspartimideEntryCount] + str(rowTracker)
    sheet[cellNum] = i
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].border = Border(bottom=thick)
    AspartimideEntryCount += 1
    
#Types "Total" heading for last column.
sheet['V' + str(rowTracker)] = "Total"
sheet['V' + str(rowTracker)].alignment = center
sheet['V' + str(rowTracker)].font = Font(size=12, bold=True)
sheet['V' + str(rowTracker)].border = Border(right=thick, bottom=thick, top=thick, left=thick)

#Types aspartimide counts for each protein into approp. cells in the sheet.
rowTracker +=1
firstAspRow = rowTracker #Important for finding total number of aspartimides later in script.
for File in sorted(glob.iglob("*.txt"), key=numericalSort):
    with open(File, "U") as inFile: #U allows any format for the .txt files.
        #Writes protein name into sheet
        ProteinName = File.rstrip(".txt")
        if ProteinName.endswith("fasta"): #Removes fasta from protein name, if there
            ProteinName = ProteinName.rstrip(".fasta")
        sheet["A" + str(rowTracker)] = ProteinName
        sheet["A" + str(rowTracker)].alignment = center
        sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
        sheet["A" + str(rowTracker)].border = Border(right=thick)
        inFile.readline() #Skips first line in FASTA files.
        
        #This saves the entire protein sequence within a single variable,
        #allowing di-peptide searching.
        ProteinRead = inFile.read()
        ProteinRead = ProteinRead.replace("\n","").replace("\r","").replace(" ","").replace("\t","")
        ProteinRead = ProteinRead.upper()
        
        #Writes the number of each possible aspartimide in the sheet.
        AspartimideEntryCount = 0
        for i in AspartimideList:
            cellNum = letterList[AspartimideEntryCount] + str(rowTracker)
            sheet[cellNum] = occurrences(ProteinRead, i)
            sheet[cellNum].alignment = Alignment(horizontal="center")
            AspartimideEntryCount +=1
        
        #Writes the total number of aspartimides found in a protein into the sheet.    
        rowTotal = ("=SUM(" + letterList[0] + str(rowTracker) + ":" +
                    letterList[AspartimideEntryCount - 1] + str(rowTracker) + ")")
        sheet[letterList[AspartimideEntryCount] + str(rowTracker)] = rowTotal
        sheet[letterList[AspartimideEntryCount] + str(rowTracker)].alignment = center
        sheet[letterList[AspartimideEntryCount] + str(rowTracker)].font = (
            Font(size=12, bold=True))
        sheet[letterList[AspartimideEntryCount] + str(rowTracker)].fill = aquaFill
        sheet[letterList[AspartimideEntryCount] + str(rowTracker)].border = (
            Border(left=thick, right=thick))
    rowTracker += 1
    inFile.close()
    
#Types "Total" heading into the row after the last protein row. 
sheet["A" + str(rowTracker)] = "Total"
sheet["A" + str(rowTracker)].alignment = center
sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
sheet["A" + str(rowTracker)].border = Border(top=thick, right=thick, bottom=thick)

#Puts total aspartimide count for each type of aspartimide into the sheet.
AspartimideEntryCount = 0
for i in range(0,21):
    cellNum = letterList[AspartimideEntryCount] + str(rowTracker)
    sheet[cellNum] = ("=SUM(" + letterList[AspartimideEntryCount] + str(firstAspRow) + ":" +
                      letterList[AspartimideEntryCount] + str(rowTracker - 1) + ")")
    sheet[cellNum].alignment = center
    sheet[cellNum].font = Font(size=12, bold=True)
    sheet[cellNum].fill = aquaFill
    sheet[cellNum].border = Border(top=thick, bottom=thick, right=thick)
    AspartimideEntryCount += 1
rowTracker += 2 #Skips a space to put in the pseudoproline info into the sheet.

#Writes initial pseudoproline information into the CPS Di-AA sheet.
#This has a different method for writing data into columns, as there are 40 possible
#pseudoprolines!
sheet.merge_cells(start_row=rowTracker, start_column=2, end_row=rowTracker,
				  end_column=(len(PSList) + 2))
sheet['B' + str(rowTracker)] = "Possible Pseudoprolines"
sheet['B' + str(rowTracker)].alignment = center
sheet['B' + str(rowTracker)].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet['B' + str(rowTracker)].fill = blackFill
rowTracker += 1

sheet['A' + str(rowTracker)] = "Protein Name"
sheet['A' + str(rowTracker)].alignment = center
sheet['A' + str(rowTracker)].font = Font(size=12, bold=True)
sheet['A' + str(rowTracker)].border = Border(right=thick, bottom=thick, top=thick)

#Writes the types of possible pseudoprolines into the top of the columns.
columnTracker = 2
for i in PSList:
    cell = sheet.cell(row = rowTracker, column = columnTracker)
    cell.value = i
    cell.alignment = center
    cell.font = Font(size=12, bold=True)
    cell.border = Border(bottom=thick)
    columnTracker += 1
    
#Types "Total" heading for last column.
cell = sheet.cell(row = rowTracker, column = columnTracker)
cell.value = "Total"
cell.alignment = center
cell.font = Font(size=12, bold=True)
cell.border = Border(right=thick, bottom=thick, top=thick, left=thick)

#Types pseudoproline counts for each protein into approp. cells in the sheet.
rowTracker +=1
firstPSRow = rowTracker #Important for finding total pseudoprolines later in script.
for File in sorted(glob.iglob("*.txt"), key=numericalSort):
    with open(File, "U") as inFile: #U allows any format for the .txt files
        #Writes protein name into sheet
        ProteinName = File.rstrip(".txt")
        if ProteinName.endswith("fasta"): #Removes fasta from protein name, if there.
            ProteinName = ProteinName.rstrip(".fasta")
        sheet["A" + str(rowTracker)] = ProteinName
        sheet["A" + str(rowTracker)].alignment = center
        sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
        sheet["A" + str(rowTracker)].border = Border(right=thick)
        inFile.readline() #Skips the first line in FASTA files
        
        #This saves the entire protein sequence within a single variable,
        #allowing di-peptide searching.
        ProteinRead = inFile.read()
        ProteinRead = ProteinRead.replace("\n","").replace("\r","").replace(" ","").replace("\t","")
        ProteinRead = ProteinRead.upper()
        
        #This writes the number of each pseudoproline into the sheet.
        columnTracker = 2
        for i in PSList:
            cell = sheet.cell(row = rowTracker, column = columnTracker)
            cell.value = occurrences(ProteinRead, i)
            cell.alignment = Alignment(horizontal="center")
            columnTracker +=1
        
        #These allow the total number of psuedoprolines to be found.
        firstColumnCell = sheet.cell(row = rowTracker, column = 2)
        finalColumnCell = sheet.cell(row = rowTracker, column = columnTracker - 1)
        
        #This writes the total number of pseudoprolines within a protein to the sheet.
        rowTotal = ("=SUM(" + firstColumnCell.column + str(firstColumnCell.row) + ":" + 
        		   finalColumnCell.column + str(finalColumnCell.row) + ")")
        rowTotalCell = sheet.cell(row = rowTracker, column = columnTracker)
        rowTotalCell.value = rowTotal
        rowTotalCell.alignment = center
        rowTotalCell.font = (Font(size=12, bold=True))
        rowTotalCell.fill = aquaFill
        rowTotalCell.border = (Border(left=thick, right=thick))
    rowTracker += 1
    inFile.close()
    
#Types "Total" heading into the row after the last protein row.  
sheet["A" + str(rowTracker)] = "Total"
sheet["A" + str(rowTracker)].alignment = center
sheet["A" + str(rowTracker)].font = Font(size=12, bold=True)
sheet["A" + str(rowTracker)].border = Border(top=thick, right=thick, bottom=thick)

#Puts total pseudoproline count for each type of pseudoproline into the sheet.
columnTracker = 2
for i in range(0,len(PSList)+1):
	firstRowCell = sheet.cell(row = firstPSRow, column = columnTracker)
	lastRowCell = sheet.cell(row = (rowTracker - 1), column = columnTracker)
	cell = sheet.cell(row = rowTracker, column = columnTracker)
	cell.value= ("=SUM(" + firstRowCell.column + str(firstRowCell.row) + ":" + 
				lastRowCell.column + str(lastRowCell.row) + ")")
	cell.alignment = center
	cell.font = Font(size=12, bold=True)
	cell.fill = aquaFill
	cell.border = Border(top=thick, bottom=thick, right=thick)
	columnTracker += 1
	
#The following codes for writing the total Di-AA composition sheet.
sheet = outFile.get_sheet_by_name("Total Di-AA Composition")

#This loop writes the total Di-AA compositions for each protein into the sheet.
entryTracker = 0
rowNum = 1
columnNum = 3
for File in sorted(glob.iglob("*.txt"), key=numericalSort):
    with open(File, "U") as inFile: #U allows any format for the .txt files
        #Writes protein name into sheet.
        ProteinName = File.rstrip(".txt")
        if ProteinName.endswith("fasta"): #Removes fasta from protein name, if there
            ProteinName = ProteinName.rstrip(".fasta")
        inFile.readline() #Skips first line in FASTA files
        
        #Writes the protein name at the top of the total Di-AA box.
        sheet.merge_cells(start_row=rowNum, start_column=columnNum, end_row=rowNum,
				  end_column=(columnNum+19))
        sheet["C" + str(rowNum)] = ProteinName
        sheet["C" + str(rowNum)].alignment = center
        sheet["C" + str(rowNum)].font = Font(size=12, color='FFFFFFFF', bold=True)
        sheet["C" + str(rowNum)].fill = blackFill
        rowNum += 1
        
        #Writes '1st Amino Acid' title into total Di-AA box.
        sheet.merge_cells(start_row=rowNum, start_column=columnNum, end_row=rowNum,
				  end_column=(columnNum + 19))
        sheet["C" + str(rowNum)] = "1st Amino Acid"
        sheet["C" + str(rowNum)].alignment = center
        sheet["C" + str(rowNum)].font = Font(size=12, bold=True)
        sheet["C" + str(rowNum)].fill = aquaFill
        rowNum += 1
        
        #Writes '2nd Amino Acid' title into total Di-AA box.
        sheet.merge_cells(start_row=rowNum+1, start_column=columnNum-2, end_row=rowNum+20,
				  end_column=(columnNum-2))
        sheet["A" + str(rowNum+1)] = "2nd Amino Acid"
        sheet["A" + str(rowNum+1)].alignment = center
        sheet["A" + str(rowNum+1)].font = Font(size=12, bold=True)
        sheet["A" + str(rowNum+1)].fill = aquaFill
        
        #Writes each amino acid next to the '2nd Amino Acid' heading.
        rowTracker = rowNum + 1
        for i in AAList:
        	cell = sheet.cell(row = rowTracker, column = columnNum-1)
        	cell.value = i
        	cell.alignment = Alignment(horizontal="center")
        	cell.font = Font(size=12, bold=True)
        	cell.border = Border(top=thick, bottom=thick, right=thick, left=thick)
        	rowTracker +=1
        
        #Writes each amino acid underneath the '1st Amino Acid' heading.
        columnTracker = columnNum	
        for i in AAList:
        	cell = sheet.cell(row = rowNum, column = columnTracker)
        	cell.value = i
        	cell.alignment = Alignment(horizontal="center")
        	cell.font = Font(size=12, bold=True)
        	cell.border = Border(top=thick, bottom=thick, right=thick, left=thick)
        	columnTracker +=1
        rowNum += 1
        
        #This saves the entire protein sequence within a single variable,
        #allowing di-peptide searching.
        ProteinRead = inFile.read()
        ProteinRead = ProteinRead.replace("\n","").replace("\r","").replace(" ","").replace("\t","")
        ProteinRead = ProteinRead.upper()
        
        #This writes the number of each di-peptide found for the protein into the sheet.
        rowTracker = rowNum
        for i in AllDiPeptideList:
        	for j in i:
        		cell = sheet.cell(row = rowTracker, column = columnNum)
        		cell.value = occurrences(ProteinRead, j)
        		cell.alignment = Alignment(horizontal="center")
        		rowTracker += 1
        	columnNum += 1
        	rowTracker = rowNum
        
        #The following skips spaces and realigns the variables to allow another di-AA box
        #to be written for the next protein in the user's folder.
        rowNum += 21
        columnNum = 3
        entryTracker += 1

#Writes total di-AA counts for all proteins into the bottom of the sheet.
#This writes the title for the total di-AA box.
sheet.merge_cells(start_row=rowNum, start_column=columnNum, end_row=rowNum,
				  end_column=(columnNum+19))
sheet["C" + str(rowNum)] = "Total Di-Amino Acid Counts"
sheet["C" + str(rowNum)].alignment = center
sheet["C" + str(rowNum)].font = Font(size=12, color='FFFFFFFF', bold=True)
sheet["C" + str(rowNum)].fill = blackFill
rowNum += 1

#This writes the '1st Amino Acid' title into the total Di-AA box.
sheet.merge_cells(start_row=rowNum, start_column=columnNum, end_row=rowNum,
		  end_column=(columnNum + 19))
sheet["C" + str(rowNum)] = "1st Amino Acid"
sheet["C" + str(rowNum)].alignment = center
sheet["C" + str(rowNum)].font = Font(size=12, bold=True)
sheet["C" + str(rowNum)].fill = aquaFill
rowNum += 1

#This writes the '2nd Amino Acid' title into the total Di-AA box.
sheet.merge_cells(start_row=rowNum+1, start_column=columnNum-2, end_row=rowNum+20,
		  end_column=(columnNum-2))
sheet["A" + str(rowNum+1)] = "2nd Amino Acid"
sheet["A" + str(rowNum+1)].alignment = center
sheet["A" + str(rowNum+1)].font = Font(size=12, bold=True)
sheet["A" + str(rowNum+1)].fill = aquaFill

#This writes each amino acid next to the '2nd Amino Acid' title.
rowTracker = rowNum + 1
for i in AAList:
	cell = sheet.cell(row = rowTracker, column = columnNum-1)
	cell.value = i
	cell.alignment = Alignment(horizontal="center")
	cell.font = Font(size=12, bold=True)
	cell.border = Border(top=thick, bottom=thick, right=thick, left=thick)
	rowTracker +=1

#This writes each amino acid underneath the '1st Amino Acid' title.
columnTracker = columnNum	
for i in AAList:
	cell = sheet.cell(row = rowNum, column = columnTracker)
	cell.value = i
	cell.alignment = Alignment(horizontal="center")
	cell.font = Font(size=12, bold=True)
	cell.border = Border(top=thick, bottom=thick, right=thick, left=thick)
	columnTracker +=1
rowNum += 1

#The following loop calculates the total of each di-AA sequence found for the entire
#set of proteins.
letterList = list(ascii_uppercase)
rowTracker = rowNum
letterCount = 2
for i in AllDiPeptideList:
	for j in i:
		letter = letterList[letterCount]
		Entry = ""
		for i in range(1,entryTracker+1):
			Entry = Entry+","+letter+str((rowTracker-(i*24)))
		cell = sheet.cell(row = rowTracker, column = columnNum)
		
		cell.value = "=SUM("+Entry+")"
		cell.alignment = Alignment(horizontal="center")
		rowTracker += 1
	columnNum += 1
	rowTracker = rowNum
	letterCount += 1

#Saves output Excel sheet based on the user's folder.
cwd = os.getcwd()
folder = os.path.basename(cwd)
outFile.save("AA Analysis for " + folder + ".xlsx")

#Concluding message to the user.
print "Paacman has finished!"